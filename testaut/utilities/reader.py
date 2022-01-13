import softest
import inspect
import logging
import pytest
import unittest
from openpyxl import Workbook, load_workbook
import xlrd
import csv

class Reader(softest.TestCase):

    def __init__(self):
        pass

    def read_data_from_excel(self,fil,sheet):
        data_list = []
        wb = load_workbook(fil)
        print(wb.sheetnames)
        sh = wb[sheet]
        row_ct = sh.max_row
        col_ct = sh.max_column
        for i in range(2, row_ct + 1):
            row = []
            for j in range(1, col_ct + 1):
                row.append(sh.cell(row=i, column=j).value)
            data_list.append(row)
        print(data_list)
        return data_list


    def read_data_from_csv(self,fil):
        with open(fil,encoding='utf8') as f:
            csvreader = csv.reader(f)
            header = next(csvreader)
            rows = []
            for row in csvreader:
                rows.append(row)
        return rows

