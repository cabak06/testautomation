import datetime
from Module1 import module1 as m1
from openpyxl import Workbook, load_workbook

current = datetime.datetime.today()

fmat = current.strftime('%Y-%m-%d-%H-%M-%S-%f')

#print(fmat)
#print('fmat is of type: ', type(fmat))


#m1.func1()
#m1.func2()

#wb = Workbook()
#ws = wb.active

#testdata = [['Name','City'],['Manish', 'Melboure'],['Rama','Bangalore'],['Sam','London']]

#for data in testdata:
#    ws.append(data)

#wb.save("demoexcel.xlsx")


#wbb = load_workbook('demoexcel.xlsx')
#shh = wbb.active
#print(shh['A3'].value)

lst = ["10.45\n"," - \n", "19.30\n","Pegasus\n","6 timer 45 minutter\n","CPH\n","–ESB\n","1 mellemlanding\n","2 timer 35 minutter SAW\n","275 kg CO2\n","Gns. udledning\n","2.030 kr.\n","returrejse\n" ]


money = "3075 kr"
aa = money.split(" ")

mon = '20.228 kr.'

print(mon[-3:])

"""

10.45
 –
19.30
Pegasus
6 timer 45 minutter
CPH
–ESB
1 mellemlanding
2 timer 35 minutter SAW
275 kg CO2
Gns. udledning
2.030 kr.
returrejse
"""